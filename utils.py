from cmath import exp
from ctypes import Union
from gc import get_objects
from itertools import count, product
from operator import eq
from os import P_ALL
from types import NoneType
from tienda.product_selection.types import ExpirationHandlingStrategy

from tienda.trade_items.selectors import SupplierProductTradeItems



from tienda.core.models import Region
from tienda.products.models import Product
from tienda.product_selection.models import ProductRegionStatus
from tienda.products.types import ProductStatus
import pandas as pd
import numpy as np
import openpyxl
import xlrd
from tienda.product_selection.models import ExpiryHandling
from tienda.suppliers.models import SupplierProduct
from tienda.product_selection.models import ProductRegionStatus
from tienda.trade_items.models import TradeItem

def product_status_id_to_name(
    product_status_id: int | str | None,
) -> str:
    status = (
        ProductStatus(int(product_status_id))
        if product_status_id and product_status_id != "None"
        else None
    )
    if status is None:
        return "Not available in region"

    return status.name


def product_has_region_status_in_country(product_id: int, country: str) -> bool:
    return ProductRegionStatus.objects.filter(
        product_id=product_id, region__site__country=country
    ).exists()


def  expiry_data_insert ()-> ExpiryHandling :
    xlrd.xlsx.ensure_elementtree_imported(False, None)
    xlrd.xlsx.Element_has_iter = True
    book = xlrd.open_workbook('/Users/kumkum/tienda/Expiry-Upload.xlsx')
    sheet = book.sheet_by_name('Sheet1')
    for r in range(1, sheet.nrows):
        exp = ExpiryHandling.objects.create(id=r,c_productid=sheet.cell(r,0).value,s_productid=sheet.cell(r,1).value,for_supplier=sheet.cell(r,2).value,for_sale=sheet.cell(r,3).value,exp_handling=sheet.cell(r,4).value)
        exp.save
    return exp

#expiry_data_insert()

def expiry_data_update_product () ->ExpiryHandling:
    updated = 0
    skipped = 0
    matched = 0
    queryset = []
    for d in ExpiryHandling.objects.all() :
        d.c_productid = str(d.c_productid).zfill(14)
        d.s_productid=  str(d.s_productid).zfill(14)
        supplier_products = SupplierProduct.objects.filter(
        price_trade_item__gtin=d.c_productid,
        trade_item__gtin=d.s_productid,
        supplier__jurisdiction__slug="germany")
        queryset1 = ProductRegionStatus.objects.filter(regiod_id=35).select_related('product').get( product_id=supplier_products.product_id)
        queryset = supplier_products
        #queryset = queryset1
        #print ( "matching record", supplier_products.all())
        matched+=supplier_products.count()
        #matched+=queryset1.count()
        #print ( "matching record", queryset1.all())

        #if supplier_products.count() == 0:
        #   print("Missing:", d.c_productid, d.c_productid)
        #  continue

        for supplier_product in supplier_products:
            #print ( "old C_GTIN is ",supplier_product.price_trade_item_id)
            #print ( "old S_GTIN is", supplier_product.trade_item_id )
            #print ( "old expiration days",supplier_product.promised_expiration_days)
            #print ( "new expiration days",d.for_supplier)
            if supplier_product.promised_expiration_days != int( d.for_supplier):
                supplier_product.promised_expiration_days = int( d.for_supplier)
                supplier_product.save()
                updated += 1
            else:
                skipped += 1

    print("Skipped:", skipped)
    print("Updated:", updated)
    
    if matched==0:
        print ( "no matching GTIN found to update the records")
    else:
        print ( "total number of matching records:", +matched)
        #print ( "matching records:", queryset.all)

#expiry_data_update_product()


def  expiry_data_update_region_status ()-> ExpiryHandling:
     updated = 0
     skipped = 0
     missing = 0
  
     for d in ExpiryHandling.objects.all():
         d.c_productid = str(d.c_productid).zfill(14)
         d.s_productid=  str(d.s_productid).zfill(14)
         if d.exp_handling.lower() == "none":
             expiration = ExpirationHandlingStrategy.NONE
         if d.exp_handling.lower() == "normal":
             expiration = ExpirationHandlingStrategy.NORMAL
         if d.exp_handling.lower() == "strict":
             expiration = ExpirationHandlingStrategy.STRICT

         try:
             trade_item = TradeItem.objects.get(gtin=d.c_productid)
         except TradeItem.DoesNotExist:
             missing += 1
             print("Could not find", d.c_productid)
             continue

         product = trade_item.product
         if not product:
             print("No product connected to", d.c_productid)
             missing += 1
             continue

         region_statuses = product.region_statuses.filter(region__jurisdiction__slug="germany")
         if region_statuses.count() == 0:
             print("Missing purchasing", d.c_productid)
             missing += 1
             continue

         for status in region_statuses:
                 should_save = False
                 print("Updating", d.c_productid, status.region)

                 if status.promised_expiration_days_for_sale != int(d.for_sale):
                   print("\t", f"promised: old={status.promised_expiration_days_for_sale}", f"new={d.for_sale}")
                   status.promised_expiration_days_for_sale = int(d.for_sale)
                   should_save = True

                 if status.expiration_handling != expiration:
                   print("\t", f"handling: old={status.expiration_handling}", f"new={expiration}")
                   status.expiration_handling = expiration
                   should_save = True

                 if should_save:
                   print("\t", "saved")
                   status.save()
                   updated += 1
                 else:
                   print("\t", "no-change")
                   skipped += 1

     print("Missing:", missing)
     print("Skipped:", skipped)
     print("Updated:", updated)
    
expiry_data_update_region_status ()